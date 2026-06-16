#!/usr/bin/env python3
"""Build a global overview of the paper-collection run.

Joins the cached pipeline state (keyword candidates + gate + enrichment + deep
analysis + archived PDFs) into:

  1. papers/<VENUE>/<readable-title>.pdf  - downloaded PDFs, renamed & grouped
                                            by venue so they are browsable in git.
  2. data/output/overview.xlsx            - one sheet per venue listing EVERY
                                            keyword candidate and its status
                                            (downloaded / irrelevant / relevant
                                            but not downloaded), plus an overview
                                            sheet with per-venue counts.

Pure offline: reuses data/raw (DBLP TOC) and data/cache (LLM results). Re-runnable.

    python3 build_overview.py
"""
import os
import re
import shutil

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
import dblp_fetch
import enrich as enrich_mod
from analyze import _safe

PDF_DIR = os.path.join(config.CACHE_DIR, "pdf")
TEXT_DIR = os.path.join(config.CACHE_DIR, "fulltext")
PAPERS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "papers")
XLSX_PATH = os.path.join(config.OUTPUT_DIR, "overview.xlsx")

# status buckets ------------------------------------------------------------
S_DOWNLOADED = "✅ 相关·已下载全文"
S_ABSTRACT = "📃 相关·仅摘要(付费墙未下载)"
S_RELEVANT_NOPDF = "⚠️ 相关·全文未归档"
S_IRRELEVANT = "❌ 不相关(初筛假阳性)"

import json


def _load(path):
    if os.path.exists(path):
        with open(path, encoding="utf-8") as fh:
            return json.load(fh)
    return None


def _safe_filename(title, maxlen=90):
    name = re.sub(r'[\\/:*?"<>|\n\r\t]', " ", title or "untitled")
    name = re.sub(r"\s+", " ", name).strip()
    return name[:maxlen].rstrip(" .")


def build_records():
    """Reconstruct every keyword candidate with its full pipeline status."""
    by_vy = dblp_fetch.fetch_all(force=False)
    candidates = enrich_mod.select_candidates(by_vy)

    records = []
    for c in candidates:
        key = c["key"]
        safe = _safe(key)
        gate = _load(os.path.join(config.CACHE_DIR, f"gate_{safe}.json")) or {}
        enr = _load(os.path.join(config.CACHE_DIR, f"enrich_{safe}.json")) or {}
        deep = _load(os.path.join(config.CACHE_DIR, f"deep_{safe}.json")) or {}
        pdf_path = os.path.join(PDF_DIR, f"{safe}.pdf")
        has_pdf = os.path.exists(pdf_path)
        has_text = os.path.exists(os.path.join(TEXT_DIR, f"{safe}.txt"))

        relevant = bool(gate.get("is_agent_security"))
        basis = deep.get("evidence_basis")
        if not relevant:
            status = S_IRRELEVANT
        elif has_pdf and basis == "full_text":
            status = S_DOWNLOADED
        elif basis == "abstract_only" or (relevant and not has_pdf):
            status = S_ABSTRACT
        else:
            status = S_RELEVANT_NOPDF

        # open-source
        os_val = str(deep.get("open_source")).lower()
        if os_val in ("true", "yes"):
            opensrc = "✅ " + (deep.get("code_url") or enr.get("github") or "")
        elif os_val in ("false", "no"):
            opensrc = "❌"
        else:
            opensrc = "未知" if relevant else ""

        records.append({
            "key": key, "safe": safe,
            "venue": c["venue"], "year": c["year"], "title": c["title"],
            "kw_priority": c.get("kw_priority", ""),
            "kw_hits": ", ".join(c.get("kw_hits", []) or []),
            "relevant": "✓" if relevant else "✗",
            "gate_conf": gate.get("confidence", ""),
            "gate_reason": gate.get("agent_relevance", ""),
            "subcategory": deep.get("subcategory") or gate.get("subcategory") or "",
            "basis": {"full_text": "全文", "abstract_only": "仅摘要"}.get(basis, ""),
            "deep_conf": deep.get("confidence", ""),
            "has_pdf": "✓" if has_pdf else "✗",
            "has_text": "✓" if has_text else "✗",
            "opensrc": opensrc,
            "arxiv": enr.get("arxiv_id") or "",
            "status": status,
            "dblp": c.get("url") or "",
            "pdf_path": pdf_path if has_pdf else None,
        })
    return records


def organize_pdfs(records):
    """Copy downloaded PDFs into papers/<VENUE>/<readable-title>.pdf (idempotent)."""
    copied = 0
    for r in records:
        if not r["pdf_path"]:
            continue
        vdir = os.path.join(PAPERS_DIR, r["venue"])
        os.makedirs(vdir, exist_ok=True)
        dst = os.path.join(vdir, f"{r['year']} - {_safe_filename(r['title'])}.pdf")
        r["paper_file"] = os.path.relpath(dst, os.path.dirname(os.path.abspath(__file__)))
        if not os.path.exists(dst):
            shutil.copy2(r["pdf_path"], dst)
            copied += 1
    return copied


# ---------------------------------------------------------------- xlsx render
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF")
STATUS_FILL = {
    S_DOWNLOADED: PatternFill("solid", fgColor="C6EFCE"),
    S_ABSTRACT: PatternFill("solid", fgColor="FFEB9C"),
    S_RELEVANT_NOPDF: PatternFill("solid", fgColor="FFD9B3"),
    S_IRRELEVANT: PatternFill("solid", fgColor="F2F2F2"),
}

COLUMNS = [
    ("#", "idx", 5),
    ("状态", "status", 24),
    ("标题", "title", 60),
    ("年份", "year", 6),
    ("子类", "subcategory", 16),
    ("相关", "relevant", 6),
    ("初筛优先级", "kw_priority", 10),
    ("命中关键词", "kw_hits", 22),
    ("Gate置信度", "gate_conf", 9),
    ("Gate判定理由", "gate_reason", 50),
    ("分析依据", "basis", 8),
    ("PDF", "has_pdf", 6),
    ("全文文本", "has_text", 8),
    ("是否开源", "opensrc", 32),
    ("arXiv", "arxiv", 14),
    ("归档路径", "paper_file", 40),
    ("DBLP", "dblp", 30),
]


def _style_header(ws, ncol):
    for j in range(1, ncol + 1):
        cell = ws.cell(row=1, column=j)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}1"
    ws.row_dimensions[1].height = 28


def _write_venue_sheet(wb, venue, recs):
    ws = wb.create_sheet(venue)
    for j, (head, _, width) in enumerate(COLUMNS, 1):
        ws.cell(row=1, column=j, value=head)
        ws.column_dimensions[get_column_letter(j)].width = width
    _style_header(ws, len(COLUMNS))
    # downloaded first, then abstract-only, then irrelevant; by year desc, title
    order = {S_DOWNLOADED: 0, S_ABSTRACT: 1, S_RELEVANT_NOPDF: 2, S_IRRELEVANT: 3}
    recs = sorted(recs, key=lambda r: (order.get(r["status"], 9), -int(r["year"]), r["title"].lower()))
    for i, r in enumerate(recs, 1):
        r["idx"] = i
        row = i + 1
        for j, (_, field, _) in enumerate(COLUMNS, 1):
            ws.cell(row=row, column=j, value=r.get(field, ""))
        fill = STATUS_FILL.get(r["status"])
        if fill:
            ws.cell(row=row, column=2).fill = fill
        ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=10).alignment = Alignment(wrap_text=True, vertical="top")


def _write_overview(wb, records):
    ws = wb.create_sheet("总览", 0)
    ws["A1"] = "AI Agent 安全论文采集 — 全局视图"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "数据来源: DBLP 四大安全顶会 (NDSS / CCS / USENIX / S&P), 2025–2026"
    ws["A3"] = ("流程: 关键词初筛 → LLM 相关性 Gate → 全文下载 → 全文深度分析(中文)。"
                "下表按会议统计每个候选的最终状态。")

    venues = ["NDSS", "CCS", "USENIX", "SP"]
    statuses = [S_DOWNLOADED, S_ABSTRACT, S_RELEVANT_NOPDF, S_IRRELEVANT]
    head = ["会议"] + statuses + ["候选合计"]
    r0 = 5
    for j, h in enumerate(head, 1):
        c = ws.cell(row=r0, column=j, value=h)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[r0].height = 30

    totals = {s: 0 for s in statuses}
    for vi, v in enumerate(venues, 1):
        recs = [r for r in records if r["venue"] == v]
        row = r0 + vi
        ws.cell(row=row, column=1, value=v).font = Font(bold=True)
        for sj, s in enumerate(statuses, 2):
            n = sum(1 for r in recs if r["status"] == s)
            totals[s] += n
            cell = ws.cell(row=row, column=sj, value=n)
            cell.alignment = Alignment(horizontal="center")
            if STATUS_FILL.get(s):
                cell.fill = STATUS_FILL[s]
        ws.cell(row=row, column=len(head), value=len(recs)).alignment = Alignment(horizontal="center")
    # totals row
    trow = r0 + len(venues) + 1
    ws.cell(row=trow, column=1, value="合计").font = Font(bold=True)
    for sj, s in enumerate(statuses, 2):
        ws.cell(row=trow, column=sj, value=totals[s]).font = Font(bold=True)
        ws.cell(row=trow, column=sj).alignment = Alignment(horizontal="center")
    ws.cell(row=trow, column=len(head), value=sum(totals.values())).font = Font(bold=True)
    ws.cell(row=trow, column=len(head)).alignment = Alignment(horizontal="center")

    # legend
    lr = trow + 2
    ws.cell(row=lr, column=1, value="状态说明").font = Font(bold=True)
    legend = [
        (S_DOWNLOADED, "Gate 判为相关，已下载开放获取 PDF 并基于全文做中文深度分析。"),
        (S_ABSTRACT, "Gate 判为相关，但全文在付费墙后(IEEE/ACM)且无 arXiv 预印本，仅基于摘要分析。"),
        (S_RELEVANT_NOPDF, "Gate 判为相关，有全文文本但 PDF 未单独归档(极少数边界情况)。"),
        (S_IRRELEVANT, "标题命中关键词进入初筛，但 LLM Gate 判定并非 agent/LLM 安全研究(假阳性)。"),
    ]
    for k, (s, desc) in enumerate(legend, 1):
        ws.cell(row=lr + k, column=1, value=s).fill = STATUS_FILL.get(s)
        ws.cell(row=lr + k, column=2, value=desc)
    ws.column_dimensions["A"].width = 30
    for col in "BCDEF":
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["B"].width = 80 if False else 16


def main():
    records = build_records()
    copied = organize_pdfs(records)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop default sheet
    _write_overview(wb, records)
    for v in ["NDSS", "CCS", "USENIX", "SP"]:
        _write_venue_sheet(wb, v, [r for r in records if r["venue"] == v])
    wb.save(XLSX_PATH)

    n = len(records)
    by_status = {}
    for r in records:
        by_status[r["status"]] = by_status.get(r["status"], 0) + 1
    print(f"candidates: {n}")
    for s, c in sorted(by_status.items()):
        print(f"  {s}: {c}")
    print(f"PDFs organized into papers/ (newly copied: {copied})")
    print(f"overview: {XLSX_PATH}")


if __name__ == "__main__":
    main()
